import sqlite3
import os
from typing import *
import itertools as it

from openpyxl import load_workbook


def main(xl_file: str, db_file: str):
    """
    main() will populate the sqlite3 database `db_file` with the relevant data from the 
    excel file `xl_file`
    """
    # Open the excel file
    workbook = load_workbook(xl_file, read_only=True)

    # Get the names of the sheets of departments
    dep_sheet_names = get_names_of_departments(workbook.sheetnames)

    # Read in all the sheets for departments into one large set of dictionaries
    all_actions = list(
        it.chain.from_iterable(
            get_row_dicts(sheet) for sheet in workbook if sheet.title in dep_sheet_names
        )
    )

    # For each unique contractor, create a unique ID
    contractors_for_db = dict(
        zip(it.count(), set(x["Global Vendor Name"] for x in all_actions))
    )

    # Create a reverse version for easy lookup of the contractor_id from vendor name
    contractors = {
        global_vendor_name: contractor_id
        for contractor_id, global_vendor_name in contractors_for_db.items()
    }

    # Add action id and contractor id to all_actions
    for idx, action in enumerate(all_actions):
        action["id"] = idx
        action["contractor_id"] = contractors[action["Global Vendor Name"]]

    # Open up the database connection
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Add the contractors to the database
    cursor.executemany(
        "INSERT INTO contractors VALUES (?,?)", contractors_for_db.items()
    )

    # Put just the items that are going to the db in a collection.
    # Those are: id, deparment, actions, dollars, contractor_id
    action_db_items = (
        (
            row["id"],
            row["department"],
            row["Number of Actions"],
            row["Dollars Obligated"],
            row["contractor_id"],
        )
        for row in all_actions
    )

    # For each action, add it to the database
    cursor.executemany("INSERT INTO actions VALUES (?,?,?,?,?)", action_db_items)

    # Commit and close the databas
    conn.commit()
    conn.close()

    return action_db_items


def get_names_of_departments(list_of_departments: List[str]) -> Set[str]:
    deps = set(n for n in list_of_departments if n.rstrip(")").endswith("00"))
    return deps


def get_row_dicts(worksheet) -> List[dict]:
    # Get the headers of the this worksheet
    headers_cells = list(worksheet.iter_rows(min_row=1, max_row=1))[0]
    headers = [h.value for h in headers_cells if h.value is not None]

    # Create a list of dictionaries
    row_dicts = [
        dict(zip(headers, [cell.value for cell in row]))
        for row in worksheet.iter_rows(min_row=2)
        if any(cell.value for cell in row)
    ]

    # Add the department name to each dictionary
    dep_name = worksheet.title
    for action in row_dicts:
        action["department"] = dep_name

    return row_dicts

